# -*- coding: utf-8 -*-
__author__ = 'XMM'

import pandas as pd
import datetime
import sqlite3 as lite
from pandas.io import sql

'''
import sys
reload(sys)
sys.setdefaultencoding('utf8')
'''


#获取前一自然日格式,每周一或者节假日后记得修改n
def get_lastday(n=-4):
    now = datetime.datetime.now()
    delta = datetime.timedelta(days=n)
    n_days = now + delta
    day = n_days.strftime('%Y%m%d')
    return day

#读取o32综合信息查询的证券组合
def return_o32_data():
    o32_file = 'D:/fixed_income/'+ get_lastday() + '/' + u'综合信息查询_组合证券' + '.xlsx'
    o32_data = pd.read_excel(o32_file,dtype={u'组合编号':str})
    o32_filter_data = o32_data[[u'日期',u'基金编号',u'基金名称',u'组合编号',u'组合名称',u'证券代码',u'证券名称',\
                                u'证券类别',u'交易市场',u'投资类型',u'净价成本',u'市值',u'当日浮动盈亏',\
                                u'总体盈亏',u'持仓多空标志']]
    return o32_filter_data

#读取固收部门的数据
def return_gs_data():
    gs_file = 'D:/fixed_income/' + get_lastday() + '/' + u'固收部数据报送' + get_lastday() + '.xlsx'
    gs_data = pd.read_excel(gs_file,header=3)
    gs_filter_data = gs_data[[u'证券名称',u'证券代码',u'证券类别',u'市场',u'求和项:净价成本',\
                              u'求和项:公允价值（净价）',u'求和项:浮动盈亏（净价）',u'求和项:资本利得',\
                              u'求和项:利息收入',u'求和项:总体盈亏']]
    gs_filter_data.columns = [u'证券名称',u'证券代码',u'证券类别',u'市场',u'净价成本',u'公允价值',u'浮动盈亏',\
                              u'资本利得',u'利息收入',u'总体盈亏']
    gs_filter_data['日期'] = get_lastday()
    gs_filter_data[u'证券类别'] = gs_filter_data[u'证券类别'].fillna(method='pad')
    gs_filter_data[u'市场'] = gs_filter_data[u'市场'].fillna(method='pad')
    return gs_filter_data

#连接sqlite数据库并将dataframe里面的数据写入数据库表中
def write_data_to_db(from_data= pd.DataFrame(),table_name = 'temptable'):
    cnx = lite.connect('data.db')
    #if_exists 預設為 failed 新建一個 Daily_Record table 並寫入 sql_df資料
    #sql(from_data, name=table_name, con=cnx)
    sql.to_sql(from_data, name=table_name, con=cnx,if_exists='replace',index=False)

#把excel数据重新整理写到新表中
def rewrite_data(db_path,exectCmd):
    conn = lite.connect(db_path)
    df = pd.read_sql_query(exectCmd, con=conn)
    return df

#比较两个表的结果并返回
def compare_data(db_path,exectCmd):
    conn = lite.connect(db_path)  # 该 API 打开一个到 SQLite 数据库文件 database 的链接，如果数据库成功打开，则返回一个连接对象
    df = pd.read_sql_query(exectCmd, con=conn)
    return df

def read_file(file_path):
    script_file = open(file_path,'r',encoding='utf-8')
    script_text = script_file.read()
    script_file.close()
    return script_text

def excelAddSheet(dataframe,excelWriter,sname = "info5"):
    from openpyxl import load_workbook
    book = load_workbook(excelWriter.path)
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter,sheet_name=sname,index=None)
    excelWriter.close()

#覆盖现有sheet
def excelCoverSheet(dataframe,excelWriter,sname = "info5"):
    from openpyxl import load_workbook
    book = load_workbook(excelWriter.path)
    if book.get_sheet_by_name(sname):
        sheet = book.get_sheet_by_name(sname)
        for row in range(1,sheet.max_row) :
            for col in range(1,sheet.max_column):
                sheet.cell(column=col, row=row, value=0)
    excelWriter.book = book
    excelWriter.sheets = dict((ws.title, ws) for ws in book.worksheets)
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sname, index=None)
    excelWriter.close()

    '''
    #book.remove_sheet(book.get_sheet_by_name(sname))
    print(book.get_sheet_by_name(sname))
    print(excelWriter.sheets.get(sname))
    print(excelWriter.curr_sheet)
    
    print(excelWriter.sheets.get(sname))
    print(excelWriter.sheets)
    print(excelWriter.sheets.values())
    print(type(excelWriter.sheets.values()))
    print(type(excelWriter.sheets.keys()))
    print(excelWriter.sheets.pop(sname))
    print(excelWriter.sheets.keys())
    '''
#调试
if __name__ == '__main__':

    o32_result = return_o32_data()
    gs_result = return_gs_data()
    write_data_to_db(o32_result,'o32_result')
    write_data_to_db(gs_result,'gs_result')

    conn = lite.connect('data.db')
    conn.execute('delete from  gs_data where gs_data.日期 = '+ get_lastday() + ';')
    sql_text = read_file('sql/gs_data.sql')
    conn.execute(sql_text)

    del_sql = "delete from o32_data where replace(o32_data.日期," + "'-'"+','+"''"+") = "  + "'" + \
              get_lastday() + "'" + ';'
    conn.execute(del_sql)
    sql_text = read_file('sql/o32_data.sql')
    conn.execute(sql_text)

    sql_text = read_file('sql/delete_o32_data_duplicate.sql')
    conn.execute(sql_text)

    conn.execute('delete from sec_type_info where  sec_type_info.日期 = ' + get_lastday() + ';')
    sql_text = read_file('sql/sec_type_info.sql')
    conn.execute(sql_text)
    conn.execute('COMMIT ')

    conn.execute('drop table sec_type_info_lastday;')
    sql_text = read_file('sql/sec_type_info_lastday.sql')
    conn.execute(sql_text)

    conn.execute('drop table sec_type_info_lastmon;')
    sql_text = read_file('sql/sec_type_info_lastmon.sql')
    conn.execute(sql_text)

    conn.execute('drop table sec_type_info_lastyear;')
    sql_text = read_file('sql/sec_type_info_lastyear.sql')
    conn.execute(sql_text)

    conn.execute('drop table sec_type_info_all;')
    sql_text = read_file('sql/sec_type_info_all.sql')
    conn.execute(sql_text)

    sql_text = read_file('sql/security_cost.sql')
    db_path = 'data.db'
    rows = compare_data(db_path, sql_text)

    rows.to_excel('D:/fixed_income/' + get_lastday() + '/' + 'cost' + get_lastday() + '.xlsx', \
                  sheet_name='cost',encoding='gb2312')
    oric_excel_file = 'D:/fixed_income/' + get_lastday() + '/固定收益监控模板搭建' + get_lastday()[4:] + \
                      '（核对版）.xlsx'
    excelWriter = pd.ExcelWriter(oric_excel_file, engine='openpyxl')
    #excelAddSheet(rows,excelWriter,'cost')
    excelCoverSheet(rows, excelWriter, 'cost')







