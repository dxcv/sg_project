# -*- coding: utf-8 -*-
__author__ = 'XMM'

import pandas as pd
import datetime
import sqlite3 as lite
from pandas.io import sql

'''
import sys
reload(sys)
sys.setdefaultencoding('utf8')
'''


#获取前一自然日格式,每周一或者节假日后记得修改n
def get_lastday(n=-2):
    now = datetime.datetime.now()
    delta = datetime.timedelta(days=n)
    n_days = now + delta
    day = n_days.strftime('%Y%m%d')
    return day

def juage_suffix(in_DataFrame = pd.DataFrame()):
    print(in_DataFrame.head(1))
    if in_DataFrame[u'市场'] == '上交所':
        return '.SH'
    elif in_DataFrame[u'市场'] == '深交所':
        return '.SZ'
    else:
        return '.IB'

def run_wsd(in_DataFrame = pd.DataFrame()):
    from WindPy import w
    w.start();
    gs_result_wind = pd.DataFrame()
    temp_result = in_DataFrame
    for code in temp_result['full_code']:
        # 债券最新面值
        latest_par = w.wsd(code, "latestpar", "ED-0D", get_lastday(), "")
        gs_result_code = latest_par.Codes[0]
        code_latest_par = latest_par.Data[0][0]
        # wind债券一级分类
        windl1_type = w.wss(code, "windl1type")
        code_windl1_type = windl1_type.Data[0][0]
        # 交易场所
        exch_city = w.wss(code, "exch_city")
        code_exch_city = exch_city.Data[0][0]
        # 估价全价（中债）
        dirty_cnbd = w.wsd(code, "dirty_cnbd", "ED0D", get_lastday(), "credibility=1")
        code_dirty_cnbd = dirty_cnbd.Data[0][0]
        # 收盘价（全价）
        dirty_price = w.wsd(code, "dirtyprice", "ED0D", get_lastday(), "")
        code_dirty_price = dirty_price.Data[0][0]

        s2 = pd.Series(
            [get_lastday(), gs_result_code, code_latest_par, code_windl1_type, code_exch_city, code_dirty_cnbd,
             code_dirty_price], \
            index=['date', 'full_code', 'code_latest_par', 'code_windl1_type', 'code_exch_city', 'code_dirty_cnbd', \
                   'code_dirty_price'])
        gs_result_wind = gs_result_wind.append(s2, ignore_index=True)
    merge_result = pd.merge(temp_result, gs_result_wind)
    w.close()
    return merge_result

#读取o32综合信息查询的证券组合
def return_o32_data():
    o32_file = 'D:/fixed_income/'+ get_lastday() + '/' + u'综合信息查询_组合证券' + '.xlsx'
    o32_data = pd.read_excel(o32_file,dtype={u'组合编号':str})
    o32_filter_data = o32_data[[u'日期',u'基金编号',u'基金名称',u'组合编号',u'组合名称',u'证券代码',u'证券名称',\
                                u'证券类别',u'交易市场',u'投资类型',u'净价成本',u'市值',u'当日浮动盈亏',\
                                u'总体盈亏',u'持仓多空标志',u'持仓']]
    return o32_filter_data

#读取固收部门的数据
def return_gs_data():
    gs_file = 'D:/fixed_income/' + get_lastday() + '/' + u'固收部数据报送' + get_lastday() + '.xlsx'
    gs_data = pd.read_excel(gs_file,header=3)
    gs_filter_data = gs_data[[u'证券名称',u'证券代码',u'证券类别',u'市场',u'求和项:持仓数量',u'求和项:净价成本',\
                              u'求和项:公允价值（净价）',u'求和项:浮动盈亏（净价）',u'求和项:资本利得',\
                              u'求和项:利息收入',u'求和项:总体盈亏']]
    gs_filter_data.columns = [u'证券名称',u'证券代码',u'证券类别',u'市场',u'持仓数量',u'净价成本',u'公允价值',u'浮动盈亏',\
                              u'资本利得',u'利息收入',u'总体盈亏']
    gs_filter_data['日期'] = get_lastday()
    gs_filter_data[u'证券类别'] = gs_filter_data[u'证券类别'].fillna(method='pad')
    gs_filter_data[u'市场'] = gs_filter_data[u'市场'].fillna(method='pad')
    return gs_filter_data

def return_gs_agent_data():
    gs_file = 'D:/fixed_income/' + get_lastday() + '/' + u'固收部数据报送' + get_lastday() + '.xlsx'
    gs_data = pd.read_excel(gs_file,sheetname=1,header=4)
    gs_filter_data = gs_data[[u'交易对手',u'证券代码',u'证券名称',u'代持到期日',u'净价价格',\
                              u'求和项:持仓数量',u'求和项:净价成本',u'求和项:公允价值(净价）',\
                              u'求和项:公允价值变动（净价）']]
    gs_filter_data.columns = [u'交易对手',u'证券代码',u'证券名称',u'代持到期日',u'净价价格',u'持仓数量',u'净价成本',\
                              u'公允价值',u'公允价值变动']
    gs_filter_data['日期'] = get_lastday()
    gs_filter_data[u'交易对手'] = gs_filter_data[u'交易对手'].fillna(method='pad')
    gs_filter_data[u'证券代码'] = gs_filter_data[u'证券代码'].fillna(method='pad')
    gs_filter_data[u'证券名称'] = gs_filter_data[u'证券名称'].fillna(method='pad')
    return gs_filter_data


#连接sqlite数据库并将dataframe里面的数据写入数据库表中
def write_data_to_db(from_data= pd.DataFrame(),table_name = 'temptable'):
    cnx = lite.connect('data.db')
    #if_exists 預設為 failed 新建一個 Daily_Record table 並寫入 sql_df資料
    #sql(from_data, name=table_name, con=cnx)
    sql.to_sql(from_data, name=table_name, con=cnx,if_exists='replace',index=False)

#把excel数据重新整理写到新表中
def rewrite_data(db_path,exectCmd):
    conn = lite.connect(db_path)
    df = pd.read_sql_query(exectCmd, con=conn)
    return df

#比较两个表的结果并返回
def compare_data(db_path,exectCmd):
    conn = lite.connect(db_path)  # 该 API 打开一个到 SQLite 数据库文件 database 的链接，如果数据库成功打开，则返回一个连接对象
    df = pd.read_sql_query(exectCmd, con=conn)
    return df

def read_file(file_path):
    script_file = open(file_path,'r',encoding='utf-8')
    script_text = script_file.read()
    script_file.close()
    return script_text

def excelAddSheet(dataframe,excelWriter,sname = "info5"):
    from openpyxl import load_workbook
    book = load_workbook(excelWriter.path)
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter,sheet_name=sname,index=None)
    excelWriter.close()

#覆盖现有sheet
def excelCoverSheet(dataframe,excelWriter,sname = "info5",ref_max_row=100,ref_max_column=100):
    from openpyxl import load_workbook
    book = load_workbook(excelWriter.path)
    if book.get_sheet_by_name(sname):
        sheet = book.get_sheet_by_name(sname)
        #print(sheet.max_row)
        #print(sheet.max_column)
        for row in range(1,ref_max_row) :
            for col in range(1,ref_max_column):
                sheet.cell(column=col, row=row, value="")
    excelWriter.book = book
    excelWriter.sheets = dict((ws.title, ws) for ws in book.worksheets)
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sname, index=None)
    excelWriter.close()

    '''
    #book.remove_sheet(book.get_sheet_by_name(sname))
    print(book.get_sheet_by_name(sname))
    print(excelWriter.sheets.get(sname))
    print(excelWriter.curr_sheet)
    
    print(excelWriter.sheets.get(sname))
    print(excelWriter.sheets)
    print(excelWriter.sheets.values())
    print(type(excelWriter.sheets.values()))
    print(type(excelWriter.sheets.keys()))
    print(excelWriter.sheets.pop(sname))
    print(excelWriter.sheets.keys())
    '''
#调试
if __name__ == '__main__':

    o32_result = return_o32_data()

    gs_result = return_gs_data()
    gs_result[u'证券后缀'] = gs_result.apply(lambda r: juage_suffix(r), axis=1)
    gs_result['full_code'] = gs_result['证券代码'].astype(str) + gs_result.apply(lambda r: juage_suffix(r), \
                                                                                       axis=1)
    print('******gs_result*start******')
    print(gs_result.head(1))
    print('******gs_result*end******')

    gs_agent_result = return_gs_agent_data()
    print(gs_agent_result.head(1))
    gs_agent_result['full_code'] = gs_agent_result['证券代码']
    print('******gs_agent_result***start****')
    print(gs_agent_result.head(3))
    print('******gs_agent_result*end******')

    write_data_to_db(o32_result,'o32_result')
    #根据gs数据，拼接wind数据
    gs_wind_result = run_wsd(gs_result)
    write_data_to_db(gs_wind_result,'gs_result')

    gs_wind_result = run_wsd(gs_agent_result)
    write_data_to_db(gs_agent_result,'gs_agent_result')

    conn = lite.connect('data.db')
    conn.execute('delete from  gs_data where gs_data.日期 = '+ get_lastday() + ';')
    sql_text = read_file('sql/gs_data.sql')
    conn.execute(sql_text)

    del_sql = "delete from o32_data where replace(o32_data.日期," + "'-'"+','+"''"+") = "  + "'" + \
              get_lastday() + "'" + ';'
    #print(del_sql)
    conn.execute(del_sql)
    sql_text = read_file('sql/o32_data.sql')
    conn.execute(sql_text)
    conn.execute('COMMIT ')

    #删除o32日内重复数据
    sql_text = read_file('sql/delete_o32_data_duplicate.sql')
    conn.execute(sql_text)

    conn.execute('drop table o32_data_uniq_orig;')
    sql_text = read_file('sql/o32_data_uniq_orig.sql')
    conn.execute(sql_text)

    conn.execute('drop table o32_data_uniq;')
    sql_text = read_file('sql/o32_data_uniq.sql')
    sql_text = sql_text + get_lastday() + ') b on(a.证券名称 = b.证券名称)'
    #print(sql_text)
    conn.execute(sql_text)

    conn.execute('delete from sec_type_info;')
    sql_text = read_file('sql/sec_type_info.sql')
    conn.execute(sql_text)
    conn.execute('COMMIT ')

    conn.execute('drop table sec_type_info_lastday;')
    sql_text = read_file('sql/sec_type_info_lastday.sql')
    sql_text = sql_text + get_lastday() + ' group by a.品种)b,sec_type_info c where b.品种 = c.品种 and b.上日日期 = c.日期'
    #print(sql_text)
    conn.execute(sql_text)

    conn.execute('drop table sec_type_info_lastmon;')
    sql_text = read_file('sql/sec_type_info_lastmon.sql')
    conn.execute(sql_text)

    conn.execute('drop table sec_type_info_lastyear;')
    sql_text = read_file('sql/sec_type_info_lastyear.sql')
    conn.execute(sql_text)

    conn.execute('drop table sec_type_info_all;')
    sql_text = read_file('sql/sec_type_info_all.sql')
    sql_text = sql_text + get_lastday() + ')c left join sec_type_info_lastmon d on c.品种 = d.品种) e left join'+ \
               ' sec_type_info_lastyear f on e.品种 = f.品种'
    #print(sql_text)
    conn.execute(sql_text)

    sql_text = read_file('sql/security_cost.sql')
    db_path = 'data.db'
    rows = compare_data(db_path, sql_text)

    rows.to_excel('D:/fixed_income/' + get_lastday() + '/' + 'cost' + get_lastday() + '.xlsx', \
                  sheet_name='cost',encoding='gb2312')
    oric_excel_file = 'D:/fixed_income/' + get_lastday() + '/固定收益监控模板搭建' + get_lastday()[4:] + \
                      '（核对版）.xlsx'
    excelWriter = pd.ExcelWriter(oric_excel_file, engine='openpyxl')
    #excelAddSheet(rows,excelWriter,'cost')
    excelCoverSheet(rows, excelWriter, 'cost')

#o32信息写入'成本更新'这张工作表
    sql_text = read_file('sql/o32_result_info.sql')
    sql_text = sql_text + ' and ' +  "replace(o32_result.日期," + "'-'"+','+"''"+") = "  + "'" + get_lastday() + "'" + ';'
    db_path = 'data.db'
    rows = compare_data(db_path, sql_text)
    oric_excel_file = 'D:/fixed_income/' + get_lastday() + '/固定收益监控模板搭建' + get_lastday()[4:] + \
                      '（核对版）.xlsx'
    excelWriter = pd.ExcelWriter(oric_excel_file, engine='openpyxl')
    excelCoverSheet(rows, excelWriter, 'o32_result_info')

# o32信息写入'代持'这张工作表
    sql_text = read_file('sql/o32_result_agent_info.sql')
    sql_text = sql_text + ' and ' + "replace(o32_result.日期," + "'-'" + ',' + "''" + ") = " + "'" + get_lastday() + "'" + ';'
    db_path = 'data.db'
    rows = compare_data(db_path, sql_text)
    oric_excel_file = 'D:/fixed_income/' + get_lastday() + '/固定收益监控模板搭建' + get_lastday()[4:] + \
                      '（核对版）.xlsx'
    excelWriter = pd.ExcelWriter(oric_excel_file, engine='openpyxl')
    excelCoverSheet(rows, excelWriter, 'o32_result_agent_info')


    conn = lite.connect('data.db')
    sql_top_bond = 'select a.日期,a.证券名称,a.证券代码,a.净价成本,cast(a.公允价值 as decimal) as 公允价值,\
                    (cast(a.公允价值 as decimal)/c.公允价值之和)*100 as 占比,\
                    cast(a.资本利得 as decimal)+cast(a.利息收入 as decima) as 已实现利润 \
                    from gs_data a,(select b.日期,sum(cast(b.公允价值 as decimal)) as 公允价值之和 \
                    from gs_data b group by b.日期)c where a.日期 = c.日期 and a.日期 = ' + get_lastday() + ';'
    df = pd.read_sql_query(sql_top_bond, con=conn)
    top_bond = df.sort_values(by=['公允价值'],ascending=False).head(5)

    oric_excel_file = 'D:/fixed_income/' + get_lastday() + '/固定收益监控模板搭建' + get_lastday()[4:] + \
                      '（核对版）.xlsx'
    excelWriter = pd.ExcelWriter(oric_excel_file, engine='openpyxl')
    #excelAddSheet(rows,excelWriter,'cost')
    excelCoverSheet(top_bond, excelWriter, 'top_bond')

    #处理中间业务
    conn = lite.connect('data.db')
    conn.execute('delete from  gs_agent_data where gs_agent_data.日期 = '+ get_lastday() + ';')
    sql_text = read_file('sql/gs_agent_data.sql')
    conn.execute(sql_text)

    del_sql = "delete from o32_agent_data where replace(o32_agent_data.日期," + "'-'"+','+"''"+") = "  + "'" + \
              get_lastday() + "'" + ';'
    #print(del_sql)
    conn.execute(del_sql)
    sql_text = read_file('sql/o32_agent_data.sql')
    conn.execute(sql_text)
    conn.execute('COMMIT ')

    #删除o32中间业务日内重复数据
    sql_text = read_file('sql/delete_o32_agent_data_duplicate.sql')
    conn.execute(sql_text)

    conn.execute('drop table o32_agent_data_uniq;')
    sql_text = read_file('sql/o32_agent_data_uniq.sql')
    sql_text = sql_text + get_lastday() + ') b on(a.证券名称 = b.证券名称)'
    #print(sql_text)
    conn.execute(sql_text)

    conn.execute('delete from sec_agent_type_info;')
    sql_text = read_file('sql/sec_agent_type_info.sql')
    sql_text = sql_text + get_lastday() + ' group by gs_agent_data.日期,o32_agent_data_uniq.品种'
    #print(sql_text)
    conn.execute(sql_text)
    conn.execute('COMMIT ')

    sql_text = read_file('sql/security_agent_cost.sql')
    db_path = 'data.db'
    rows = compare_data(db_path, sql_text)

    oric_excel_file = 'D:/fixed_income/' + get_lastday() + '/固定收益监控模板搭建' + get_lastday()[4:] + \
                      '（核对版）.xlsx'
    excelWriter = pd.ExcelWriter(oric_excel_file, engine='openpyxl')
    excelCoverSheet(rows, excelWriter, 'agent_cost')

#中间业务前5大交易对手交易数据
    conn = lite.connect('data.db')
    sql_top_agent_bond = 'select b.日期,b.交易对手,b.公允价值 from \
                          (select a.日期,a.交易对手,sum(a.公允价值) as 公允价值\
                    from gs_agent_data a group by a.日期,a.交易对手) b where b.日期 =  ' + get_lastday() +\
                   ' order by b.公允价值 desc;'
    print(sql_top_agent_bond)
    df = pd.read_sql_query(sql_top_agent_bond, con=conn)
    top_agent_bond = df.sort_values(by=['公允价值'],ascending=False).head(5)

    oric_excel_file = 'D:/fixed_income/' + get_lastday() + '/固定收益监控模板搭建' + get_lastday()[4:] + \
                      '（核对版）.xlsx'
    excelWriter = pd.ExcelWriter(oric_excel_file, engine='openpyxl')
    excelCoverSheet(top_agent_bond, excelWriter, 'top_agent_bond')



