#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/3/21 21:29
# @Author  : XMM
# @Site    : 
# @File    : modify_excel.py
# @Software: PyCharm

from openpyxl import load_workbook

#print(wb.get_sheet_names())
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
#print(wb.active)

# 根据sheet名字获得sheet
#a_sheet = wb.get_sheet_by_name('期货持仓和盈亏')
#print(a_sheet.title)

# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
# 默认可读写，若有需要可以指定write_only和read_only为True
wb = load_workbook('D:/fixed_income/a.xlsx')
sheet = wb.active
b4 = sheet['B4']
print(f'({b4.column}, {b4.row}) is {b4.value}')

# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = sheet.cell(row=4, column=2)
print(b4_too.value)

# 获得最大列和最大行
print(f'max_row is {sheet.max_row}')
print(f'max column is {sheet.max_column}')

'''
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)
'''
#取表格第二列的数据
for cell in list(sheet.columns)[1]:
    print(cell.value)

for i in range(1, 4):
    for j in range(1, 3):
        print(sheet.cell(row=i, column=j))
